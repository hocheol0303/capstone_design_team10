"""
add_case.py

"맞춤솔루션출력조건_260318.xlsx"의 종합 시트에서 N개 랜덤 샘플 후:
1. 각 특성 시트에서 customer_desc 조회 → vlm_output 구성
2. 특성별 FeatureRule 벡터 검색
3. CombinedRule 노드 생성 (환자 케이스)
4. FeatureRule -[CONTRIBUTES_TO]-> CombinedRule 관계 생성
5. 최종 시술 추천(output_text) 반환

사용법:
    python add_case.py          # 기본 3개 샘플
    python add_case.py 10       # 10개 샘플
"""

import os
import random
import re
import sys
import uuid
from datetime import datetime
from pathlib import Path

from neo4j import GraphDatabase
from openai import OpenAI
from openpyxl import load_workbook


# ── 환경변수 로드 ──────────────────────────────────────────────────────────────

def load_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip().strip("'\""))


load_env_file(Path(__file__).resolve().parent / ".env")

DEFAULT_EXCEL_PATH = Path(__file__).resolve().parent / "data" / "맞춤솔루션출력조건_260318.xlsx"
EXCEL_PATH = os.getenv("EXCEL_PATH", str(DEFAULT_EXCEL_PATH))

NEO4J_URI = os.getenv("NEO4J_URI", "")
NEO4J_USERNAME = os.getenv("NEO4J_USERNAME", "")
NEO4J_PASSWORD = os.getenv("NEO4J_PASSWORD", "")
NEO4J_DATABASE = os.getenv("NEO4J_DATABASE", "").strip()
OPENAI_EMBED_MODEL = os.getenv("OPENAI_EMBED_MODEL", "text-embedding-3-small")
OPENAI_EMBED_DIMENSIONS = int(os.getenv("OPENAI_EMBED_DIMENSIONS", "1536"))

# 종합 시트의 컬럼 인덱스(0-based) → (vlm 출력에 쓸 특성명, 엑셀 시트명)
# 종합 시트: [age_group, 색소코드, 이마코드, 눈가코드, 팔자코드, 입가코드,
#             리프팅윤곽코드, 리프팅탄력코드, 스킨부스터코드, 볼꺼짐코드, ...]
FEATURE_MAP = [
    (1, "색소",            "색소"),
    (2, "이마·미간주름",   "이마·미간주름"),
    (3, "눈가·앞광대주름", "눈가·앞광대주름"),
    (4, "팔자주름",        "팔자주름"),
    (5, "입가·턱주름",     "입가·턱주름"),
    (6, "리프팅(윤곽)",    "리프팅(윤곽)"),
    (7, "리프팅(탄력)",    "리프팅(탄력)"),
    (8, "스킨부스터 (new)","스킨부스터 (new)"),
    (9, "볼꺼짐",          "볼꺼짐"),
]

openai_client = OpenAI()
driver = GraphDatabase.driver(NEO4J_URI, auth=(NEO4J_USERNAME, NEO4J_PASSWORD))


def _session():
    if NEO4J_DATABASE:
        return driver.session(database=NEO4J_DATABASE)
    return driver.session()


# ── 엑셀 유틸 ─────────────────────────────────────────────────────────────────

def clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def get_customer_desc(ws, age_group: str, code: str) -> str | None:
    """
    feature 시트(ws)에서 age_group + code가 일치하는 행의 customer_desc 반환.
    단순/다중 조건 시트 모두 지원.
    """
    headers = [clean_text(cell.value) for cell in ws[1]]
    is_multi = "1차 조건" in headers

    for row_idx in range(2, ws.max_row + 1):
        row = [clean_text(ws.cell(row_idx, c).value) for c in range(1, ws.max_column + 1)]
        if not any(row):
            continue

        if is_multi:
            # [age_group, cond1, cond2, cond3, output_text, code, customer_desc, ...]
            r_age = row[0]
            r_code = row[5] if len(row) > 5 else ""
            r_customer_desc = row[6] if len(row) > 6 else ""
        else:
            # [age_group, condition, output_text, code, customer_desc, ...]
            r_age = row[0]
            r_code = row[3] if len(row) > 3 else ""
            r_customer_desc = row[4] if len(row) > 4 else ""

        if r_age == age_group and r_code == code:
            return r_customer_desc

    return None


def sample_vlm_outputs(excel_path: str, n: int = 3) -> list[str]:
    """
    종합 시트에서 n개 랜덤 샘플 후 각각 vlm_output 문자열로 변환.
    각 특성 코드에 해당하는 customer_desc를 feature 시트에서 조회.
    """
    wb = load_workbook(excel_path, data_only=True)

    if "종합 (더 이상 사용안함)" not in wb.sheetnames:
        raise ValueError("종합 (더 이상 사용안함) 시트가 없습니다.")

    ws_combined = wb["종합 (더 이상 사용안함)"]

    # 데이터 행 수집 (빈 행 제외)
    all_rows = []
    for row_idx in range(2, ws_combined.max_row + 1):
        row = [clean_text(ws_combined.cell(row_idx, c).value) for c in range(1, 14)]
        if any(row):
            all_rows.append(row)

    if not all_rows:
        raise ValueError("종합 (더 이상 사용안함) 시트에 데이터가 없습니다.")

    sampled = random.sample(all_rows, min(n, len(all_rows)))
    vlm_outputs = []

    for row in sampled:
        age_group = row[0]
        parts = []

        for col_idx, feature_name, sheet_name in FEATURE_MAP:
            code = row[col_idx] if col_idx < len(row) else ""
            if not code or code in ("없음", "-", ""):
                continue
            if sheet_name not in wb.sheetnames:
                continue

            ws_feature = wb[sheet_name]
            desc = get_customer_desc(ws_feature, age_group, code)
            if desc:
                parts.append(f"{feature_name}: {desc}")

        if parts:
            vlm_outputs.append("; ".join(parts))

    return vlm_outputs


# ── 스키마 초기화 ──────────────────────────────────────────────────────────────

def setup_case_schema() -> None:
    with _session() as session:
        session.run(
            "CREATE CONSTRAINT combined_case_id IF NOT EXISTS "
            "FOR (n:CombinedRule) REQUIRE n.case_id IS UNIQUE"
        )


# ── 임베딩 ────────────────────────────────────────────────────────────────────

def embed(text: str) -> list:
    response = openai_client.embeddings.create(
        model=OPENAI_EMBED_MODEL,
        input=text,
        dimensions=OPENAI_EMBED_DIMENSIONS,
    )
    return response.data[0].embedding


# ── 벡터 검색 ─────────────────────────────────────────────────────────────────

def search_feature(desc: str, feature_name: str = None, top_k: int = 3) -> dict | None:
    """desc 텍스트를 임베딩하여 가장 유사한 FeatureRule 반환."""
    vector = embed(desc)
    with _session() as session:
        if feature_name:
            result = session.run(
                """
                CALL db.index.vector.queryNodes('feature_rule_embedding_idx', $top_k, $vector)
                YIELD node, score
                WHERE node.feature_name = $feature_name
                RETURN node.rule_id      AS rule_id,
                       node.feature_name AS feature_name,
                       node.output_text  AS output_text,
                       node.customer_desc AS customer_desc,
                       score
                LIMIT 1
                """,
                top_k=top_k,
                vector=vector,
                feature_name=feature_name,
            )
        else:
            result = session.run(
                """
                CALL db.index.vector.queryNodes('feature_rule_embedding_idx', 1, $vector)
                YIELD node, score
                RETURN node.rule_id      AS rule_id,
                       node.feature_name AS feature_name,
                       node.output_text  AS output_text,
                       node.customer_desc AS customer_desc,
                       score
                """,
                vector=vector,
            )
        rows = [dict(r) for r in result]
        return rows[0] if rows else None


# ── DB 쓰기 ───────────────────────────────────────────────────────────────────

def create_combined_node(case_id: str, vlm_output: str, output_text: str) -> None:
    with _session() as session:
        session.run(
            """
            MERGE (c:CombinedRule {case_id: $case_id})
            SET c.customer_desc = $customer_desc,
                c.output_text   = $output_text,
                c.created_at    = $created_at
            """,
            case_id=case_id,
            customer_desc=vlm_output,
            output_text=output_text,
            created_at=datetime.now().isoformat(),
        )


def create_contributes_to(rule_id: str, case_id: str, feature_name: str, score: float) -> None:
    with _session() as session:
        session.run(
            """
            MATCH (f:FeatureRule {rule_id: $rule_id})
            MATCH (c:CombinedRule {case_id: $case_id})
            MERGE (f)-[:CONTRIBUTES_TO {feature_name: $feature_name, score: $score}]->(c)
            """,
            rule_id=rule_id,
            case_id=case_id,
            feature_name=feature_name,
            score=score,
        )


# ── 핵심 함수 ─────────────────────────────────────────────────────────────────

def add_case(vlm_output: str) -> str:
    """VLM 출력을 DB에 케이스로 등록하고 시술 추천을 반환."""
    case_id = str(uuid.uuid4())
    matched_outputs = []
    matched_rules = []

    print(f"\n[케이스 ID] {case_id}")
    print("=" * 60)

    for block in vlm_output.split(";"):
        block = block.strip()
        if not block or ":" not in block:
            continue

        feature_name, desc = block.split(":", 1)
        feature_name = feature_name.strip()
        desc = desc.strip()

        if not desc:
            continue

        match = search_feature(desc, feature_name=feature_name)
        if match:
            print(f"  [{feature_name}]")
            print(f"    검색 결과: {match['output_text']}")
            print(f"    유사도:   {match['score']:.4f}")
            matched_outputs.append(match["output_text"])
            matched_rules.append(match)
        else:
            print(f"  [{feature_name}] → 매칭 결과 없음")

    if not matched_rules:
        print("\n매칭된 FeatureRule이 없어 케이스를 등록하지 않습니다.")
        return ""

    combined_output = "; ".join(matched_outputs)
    create_combined_node(case_id, vlm_output, combined_output)

    for match in matched_rules:
        create_contributes_to(
            rule_id=match["rule_id"],
            case_id=case_id,
            feature_name=match["feature_name"],
            score=match["score"],
        )

    print("=" * 60)
    print(f"[등록 완료] CombinedRule 노드 + {len(matched_rules)}개 관계 생성")
    return combined_output


# ── 실행 ──────────────────────────────────────────────────────────────────────

def main(n: int = 3):
    setup_case_schema()

    print(f"종합 시트에서 {n}개 랜덤 샘플 중...")
    vlm_outputs = sample_vlm_outputs(EXCEL_PATH, n=n)
    print(f"{len(vlm_outputs)}개 케이스 구성 완료\n")

    for i, vlm_output in enumerate(vlm_outputs, 1):
        print(f"\n{'━' * 60}")
        print(f"케이스 {i}/{len(vlm_outputs)}")
        print(f"{'━' * 60}")
        print("[입력]")
        print(vlm_output)
        print("\n[검색 및 등록]")
        result = add_case(vlm_output)
        print("\n[최종 시술 추천]")
        print(result)

    driver.close()


if __name__ == "__main__":
    n = int(sys.argv[1]) if len(sys.argv) > 1 else 3
    main(n)
