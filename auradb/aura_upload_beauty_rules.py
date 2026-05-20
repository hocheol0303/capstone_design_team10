import os
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openai import OpenAI, RateLimitError
from neo4j import GraphDatabase
from openpyxl import load_workbook


def load_env_file(env_path: Path) -> None:
    if not env_path.exists():
        return

    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue

        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip("'\"")
        os.environ.setdefault(key, value)


load_env_file(Path(__file__).resolve().parent.parent / ".env")

DEFAULT_EXCEL_PATH = Path(__file__).resolve().parent / "data" / "mapping_table_tmp.xlsx"
EXCEL_PATH = os.getenv("EXCEL_PATH", str(DEFAULT_EXCEL_PATH))
OPENAI_EMBED_MODEL = os.getenv("OPENAI_EMBED_MODEL", "text-embedding-3-small")
OPENAI_EMBED_DIMENSIONS = int(os.getenv("OPENAI_EMBED_DIMENSIONS", "1536"))
EMBED_BATCH_SIZE = int(os.getenv("EMBED_BATCH_SIZE", "16"))
EMBED_MIN_BATCH_SIZE = int(os.getenv("EMBED_MIN_BATCH_SIZE", "4"))
EMBED_MAX_RETRIES = int(os.getenv("EMBED_MAX_RETRIES", "8"))
EMBED_RETRY_BASE_SECONDS = float(os.getenv("EMBED_RETRY_BASE_SECONDS", "1.0"))

NEO4J_URI = os.getenv("NEO4J_URI", "")
NEO4J_USERNAME = os.getenv("NEO4J_USERNAME", "")
NEO4J_PASSWORD = os.getenv("NEO4J_PASSWORD", "")
NEO4J_DATABASE = os.getenv("NEO4J_DATABASE", "").strip()
if None in {NEO4J_URI, NEO4J_USERNAME, NEO4J_PASSWORD}:
    raise EnvironmentError("NEO4J_URI, NEO4J_USERNAME, NEO4J_PASSWORD 환경변수를 모두 설정해주세요.")
else:
    print(f"\033[42mNeo4j 연결 정보: URI={NEO4J_URI}, USERNAME={NEO4J_USERNAME}, DATABASE={NEO4J_DATABASE or '기본'}\033[0m")

FEATURE_SHEETS = [
    "색소",
    "이마·미간주름",
    "눈가·앞광대주름",
    "팔자주름",
    "입가·턱주름",
    "리프팅(윤곽)",
    "리프팅(탄력)",
    "스킨부스터 (new)",
    "볼꺼짐",
]

COMBINED_SHEET = "dummy"


@dataclass
class FeatureRule:
    rule_id: str
    feature_name: str
    age_group: str
    condition_1: Optional[str]
    condition_2: Optional[str]
    condition_3: Optional[str]
    raw_condition: str
    code: str
    output_text: str
    customer_desc: str
    clinician_desc: str
    source_sheet: str
    embedding_text: str


@dataclass
class CombinedRule:
    rule_id: str
    age_group: str
    pigment_code: str
    forehead_code: str
    eye_code: str
    nasolabial_code: str
    mouth_code: str
    contour_code: str
    tightness_code: str
    skinbooster_code: str
    volume_code: str
    output_text: str
    customer_desc: str
    clinician_desc: str
    source_sheet: str
    embedding_text: str


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def slugify(text: str) -> str:
    text = re.sub(r"\s+", "_", clean_text(text))
    text = re.sub(r"[^0-9A-Za-z가-힣_·()\-]", "", text)
    return text[:120] if text else "EMPTY"


def build_feature_embedding_text(record: FeatureRule) -> str:
    parts = [
        f"문서유형: 특성 규칙",
        f"시트명: {record.source_sheet}",
        f"특성명: {record.feature_name}",
        f"연령대: {record.age_group}",
        f"1차 조건: {record.condition_1 or '없음'}",
        f"2차 조건: {record.condition_2 or '없음'}",
        f"3차 조건: {record.condition_3 or '없음'}",
        f"원본 조건: {record.raw_condition or '없음'}",
        f"연결 코드: {record.code}",
        f"출력값: {record.output_text or '없음'}",
        f"고객용 설명: {record.customer_desc or '없음'}",
        f"의료진용 설명: {record.clinician_desc or '없음'}",
    ]
    return "\n".join(parts)


def build_combined_embedding_text(record: CombinedRule) -> str:
    parts = [
        "문서유형: 종합 규칙",
        f"시트명: {record.source_sheet}",
        f"연령대: {record.age_group}",
        f"색소 코드: {record.pigment_code}",
        f"이마·미간주름 코드: {record.forehead_code}",
        f"눈가·앞광대주름 코드: {record.eye_code}",
        f"팔자주름 코드: {record.nasolabial_code}",
        f"입가·턱주름 코드: {record.mouth_code}",
        f"리프팅(윤곽) 코드: {record.contour_code}",
        f"리프팅(탄력) 코드: {record.tightness_code}",
        f"스킨부스터 코드: {record.skinbooster_code}",
        f"볼꺼짐 코드: {record.volume_code}",
        f"출력값: {record.output_text or '없음'}",
        f"고객용 설명: {record.customer_desc or '없음'}",
        f"의료진용 설명: {record.clinician_desc or '없음'}",
    ]
    return "\n".join(parts)


def parse_feature_sheet(ws) -> List[FeatureRule]:
    headers = [clean_text(cell.value) for cell in ws[1]]
    rules: List[FeatureRule] = []

    is_three_condition = "1차 조건" in headers
    is_skinbooster_two_condition = "광채" in headers and "거칠기" in headers

    for row_idx in range(2, ws.max_row + 1):
        row = [clean_text(ws.cell(row_idx, c).value) for c in range(1, ws.max_column + 1)]
        if not any(row):
            continue

        if is_three_condition:
            # [연령, 1차, 2차, 3차, 출력값, 연결 코드, 설명(고객용), Prompt 용]
            age_group = row[0]
            condition_1 = row[1]
            condition_2 = row[2]
            condition_3 = row[3]
            output_text = row[4]
            code = row[5]
            customer_desc = row[6]
            clinician_desc = row[7] if len(row) > 7 else ""
            raw_condition = " | ".join([x for x in [condition_1, condition_2, condition_3] if x and x != "없음"])
        elif is_skinbooster_two_condition:
            # [연령, 광채, 거칠기, 출력값, 연결 코드, 설명(고객용), Prompt 용, ...]
            age_group = row[0]
            condition_1 = row[1]   # 광채
            condition_2 = row[2]   # 거칠기
            condition_3 = None
            output_text = row[3]
            code = row[4]
            customer_desc = row[5]
            clinician_desc = row[6] if len(row) > 6 else ""
            raw_condition = f"광채={condition_1} | 거칠기={condition_2}"
        else:
            # [연령, 조건값, 출력값, 연결 코드, 설명(고객용), Prompt 용]
            age_group = row[0]
            condition_1 = row[1]
            condition_2 = None
            condition_3 = None
            output_text = row[2]
            code = row[3]
            customer_desc = row[4]
            clinician_desc = row[5] if len(row) > 5 else ""
            raw_condition = condition_1

        if not output_text and not customer_desc:
            continue

        rule_id = f"feature::{slugify(ws.title)}::{slugify(age_group)}::{slugify(code)}::{row_idx}"
        record = FeatureRule(
            rule_id=rule_id,
            feature_name=ws.title,
            age_group=age_group,
            condition_1=condition_1,
            condition_2=condition_2,
            condition_3=condition_3,
            raw_condition=raw_condition,
            code=code,
            output_text=output_text,
            customer_desc=customer_desc,
            clinician_desc=clinician_desc,
            source_sheet=ws.title,
            embedding_text="",
        )
        record.embedding_text = build_feature_embedding_text(record)
        rules.append(record)

    return rules


def parse_combined_sheet(ws) -> List[CombinedRule]:
    rules: List[CombinedRule] = []

    for row_idx in range(2, ws.max_row + 1):
        row = [clean_text(ws.cell(row_idx, c).value) for c in range(1, 14)]
        if not any(row):
            continue

        age_group = row[0]
        pigment_code = row[1]
        forehead_code = row[2]
        eye_code = row[3]
        nasolabial_code = row[4]
        mouth_code = row[5]
        contour_code = row[6]
        tightness_code = row[7]
        skinbooster_code = row[8]
        volume_code = row[9]
        output_text = row[10]
        customer_desc = row[11]
        clinician_desc = row[12] if len(row) > 12 else ""

        if not age_group:
            continue

        key = "::".join([
            slugify(age_group), slugify(pigment_code), slugify(forehead_code), slugify(eye_code),
            slugify(nasolabial_code), slugify(mouth_code), slugify(contour_code),
            slugify(tightness_code), slugify(skinbooster_code), slugify(volume_code)
        ])
        rule_id = f"combined::{key}::{row_idx}"

        record = CombinedRule(
            rule_id=rule_id,
            age_group=age_group,
            pigment_code=pigment_code,
            forehead_code=forehead_code,
            eye_code=eye_code,
            nasolabial_code=nasolabial_code,
            mouth_code=mouth_code,
            contour_code=contour_code,
            tightness_code=tightness_code,
            skinbooster_code=skinbooster_code,
            volume_code=volume_code,
            output_text=output_text,
            customer_desc=customer_desc,
            clinician_desc=clinician_desc,
            source_sheet=ws.title,
            embedding_text="",
        )
        record.embedding_text = build_combined_embedding_text(record)
        rules.append(record)

    return rules

def load_rules_from_excel(excel_path: str) -> Tuple[List[FeatureRule], List[CombinedRule]]:
    wb = load_workbook(excel_path, data_only=False)

    feature_rules: List[FeatureRule] = []
    for sheet_name in FEATURE_SHEETS:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트가 없습니다: {sheet_name}")
        feature_rules.extend(parse_feature_sheet(wb[sheet_name]))

    if COMBINED_SHEET not in wb.sheetnames:
        # raise ValueError(f"시트가 없습니다: {COMBINED_SHEET}")
        print(f"경고: 시트가 없습니다: {COMBINED_SHEET} - 종합 규칙이 로드되지 않습니다")
        combined_rules = []
    else:
        combined_rules = parse_combined_sheet(wb[COMBINED_SHEET])
    return feature_rules, combined_rules


class Embedder:
    def __init__(self, model: str, dimensions: int):
        self.client = OpenAI()
        self.model = model
        self.dimensions = dimensions

    def embed_texts(self, texts: List[str], batch_size: int = 64) -> List[List[float]]:
        if not texts:
            return []

        vectors: List[List[float]] = []
        start = 0
        current_batch_size = max(1, batch_size)

        while start < len(texts):
            batch = texts[start : start + current_batch_size]
            last_error: Optional[RateLimitError] = None

            for attempt in range(EMBED_MAX_RETRIES):
                try:
                    response = self.client.embeddings.create(
                        model=self.model,
                        input=batch,
                        dimensions=self.dimensions,
                    )
                    vectors.extend([item.embedding for item in response.data])
                    start += len(batch)
                    break
                except RateLimitError as exc:
                    last_error = exc
                    wait_seconds = self._get_retry_delay(exc, attempt)

                    if len(batch) > EMBED_MIN_BATCH_SIZE:
                        current_batch_size = max(EMBED_MIN_BATCH_SIZE, len(batch) // 2)
                        batch = texts[start : start + current_batch_size]
                        print(
                            f"[embed] rate limit 감지: 배치 크기를 {len(batch)}개로 줄여 재시도합니다 "
                            f"(대기 {wait_seconds:.2f}초)"
                        )
                    else:
                        print(f"[embed] rate limit 감지: {wait_seconds:.2f}초 후 재시도합니다")

                    time.sleep(wait_seconds)
            else:
                raise RuntimeError(
                    f"임베딩 생성이 반복적으로 제한되었습니다. start={start}, "
                    f"batch_size={len(batch)}, max_retries={EMBED_MAX_RETRIES}"
                ) from last_error

        return vectors

    @staticmethod
    def _get_retry_delay(exc: RateLimitError, attempt: int) -> float:
        message = ""
        if getattr(exc, "response", None) is not None:
            try:
                payload = exc.response.json()
                message = payload.get("error", {}).get("message", "")
            except Exception:
                message = ""

        match = re.search(r"try again in (\d+)ms", message, re.IGNORECASE)
        if match:
            return max(int(match.group(1)) / 1000.0, EMBED_RETRY_BASE_SECONDS)

        return EMBED_RETRY_BASE_SECONDS * (2 ** attempt)


class AuraUploader:
    def __init__(self, uri: str, username: str, password: str, database: str):
        self.driver = GraphDatabase.driver(uri, auth=(username, password))
        self.database = database.strip()

    def _session(self):
        if self.database:
            return self.driver.session(database=self.database)
        return self.driver.session()

    def close(self) -> None:
        self.driver.close()

    def verify_connectivity(self) -> None:
        try:
            self.driver.verify_connectivity()
            with self._session() as session:
                session.run("RETURN 1 AS ok").single()
        except Exception as exc:
            raise ConnectionError(
                "Neo4j 연결에 실패했습니다. NEO4J_URI, NEO4J_USERNAME, NEO4J_PASSWORD, NEO4J_DATABASE 값을 확인해주세요."
            ) from exc

    def wipe_existing(self) -> None:
        """기존 FeatureRule/CombinedRule 노드와 벡터 인덱스를 모두 제거.

        misaligned key(예: 스킨부스터의 code 컬럼이 출력값 문자열이었던 잔여물)가
        MERGE로 덮어쓰이지 않으므로, 재적재 전에 한 번 청소한다.
        """
        with self._session() as session:
            session.run("DROP INDEX feature_rule_embedding_idx IF EXISTS")
            session.run("DROP INDEX combined_rule_embedding_idx IF EXISTS")
            session.run("MATCH (n:FeatureRule) DETACH DELETE n")
            session.run("MATCH (n:CombinedRule) DETACH DELETE n")

    def setup_schema(self, embedding_dimensions: int) -> None:
        with self._session() as session:
            session.run("CREATE CONSTRAINT feature_rule_id IF NOT EXISTS FOR (n:FeatureRule) REQUIRE n.rule_id IS UNIQUE")
            session.run("CREATE CONSTRAINT combined_rule_id IF NOT EXISTS FOR (n:CombinedRule) REQUIRE n.rule_id IS UNIQUE")
            session.run("CREATE CONSTRAINT feature_code_unique IF NOT EXISTS FOR (n:FeatureRule) REQUIRE (n.feature_name, n.age_group, n.code) IS UNIQUE")
            session.run(
                f"""
                CREATE VECTOR INDEX feature_rule_embedding_idx IF NOT EXISTS
                FOR (n:FeatureRule) ON (n.embedding)
                OPTIONS {{indexConfig: {{
                    `vector.dimensions`: {embedding_dimensions},
                    `vector.similarity_function`: 'cosine'
                }}}}
                """
            )
            session.run(
                f"""
                CREATE VECTOR INDEX combined_rule_embedding_idx IF NOT EXISTS
                FOR (n:CombinedRule) ON (n.embedding)
                OPTIONS {{indexConfig: {{
                    `vector.dimensions`: {embedding_dimensions},
                    `vector.similarity_function`: 'cosine'
                }}}}
                """
            )

    def upload_feature_rules(self, feature_rules: List[FeatureRule], vectors: List[List[float]]) -> None:
        if len(feature_rules) != len(vectors):
            raise ValueError(f"feature_rules 수({len(feature_rules)})와 임베딩 수({len(vectors)})가 다릅니다.")

        rows = []
        for rule, vector in zip(feature_rules, vectors):
            rows.append({
                "rule_id": rule.rule_id,
                "feature_name": rule.feature_name,
                "age_group": rule.age_group,
                "condition_1": rule.condition_1,
                "condition_2": rule.condition_2,
                "condition_3": rule.condition_3,
                "raw_condition": rule.raw_condition,
                "code": rule.code,
                "output_text": rule.output_text,
                "customer_desc": rule.customer_desc,
                "clinician_desc": rule.clinician_desc,
                "source_sheet": rule.source_sheet,
                "embedding_text": rule.embedding_text,
                "embedding": vector,
            })

        query = """
        UNWIND $rows AS row
        MERGE (n:FeatureRule {feature_name: row.feature_name, age_group: row.age_group, code: row.code})
        SET n.rule_id = row.rule_id,
            n.condition_1 = row.condition_1,
            n.condition_2 = row.condition_2,
            n.condition_3 = row.condition_3,
            n.raw_condition = row.raw_condition,
            n.output_text = row.output_text,
            n.customer_desc = row.customer_desc,
            n.clinician_desc = row.clinician_desc,
            n.source_sheet = row.source_sheet,
            n.embedding_text = row.embedding_text,
            n.embedding = row.embedding
        """

        with self._session() as session:
            session.run(query, rows=rows)

    def upload_combined_rules(self, combined_rules: List[CombinedRule], vectors: List[List[float]]) -> None:
        if len(combined_rules) != len(vectors):
            raise ValueError(f"combined_rules 수({len(combined_rules)})와 임베딩 수({len(vectors)})가 다릅니다.")

        rows = []
        for rule, vector in zip(combined_rules, vectors):
            rows.append({
                "rule_id": rule.rule_id,
                "age_group": rule.age_group,
                "pigment_code": rule.pigment_code,
                "forehead_code": rule.forehead_code,
                "eye_code": rule.eye_code,
                "nasolabial_code": rule.nasolabial_code,
                "mouth_code": rule.mouth_code,
                "contour_code": rule.contour_code,
                "tightness_code": rule.tightness_code,
                "skinbooster_code": rule.skinbooster_code,
                "volume_code": rule.volume_code,
                "output_text": rule.output_text,
                "customer_desc": rule.customer_desc,
                "clinician_desc": rule.clinician_desc,
                "source_sheet": rule.source_sheet,
                "embedding_text": rule.embedding_text,
                "embedding": vector,
            })

        query = """
        UNWIND $rows AS row
        MERGE (n:CombinedRule {rule_id: row.rule_id})
        SET n.age_group = row.age_group,
            n.pigment_code = row.pigment_code,
            n.forehead_code = row.forehead_code,
            n.eye_code = row.eye_code,
            n.nasolabial_code = row.nasolabial_code,
            n.mouth_code = row.mouth_code,
            n.contour_code = row.contour_code,
            n.tightness_code = row.tightness_code,
            n.skinbooster_code = row.skinbooster_code,
            n.volume_code = row.volume_code,
            n.output_text = row.output_text,
            n.customer_desc = row.customer_desc,
            n.clinician_desc = row.clinician_desc,
            n.source_sheet = row.source_sheet,
            n.embedding_text = row.embedding_text,
            n.embedding = row.embedding
        """

        with self._session() as session:
            session.run(query, rows=rows)

    def create_relationships(self) -> None:
        mappings = [
            ("pigment_code", "색소"),
            ("forehead_code", "이마·미간주름"),
            ("eye_code", "눈가·앞광대주름"),
            ("nasolabial_code", "팔자주름"),
            ("mouth_code", "입가·턱주름"),
            ("contour_code", "리프팅(윤곽)"),
            ("tightness_code", "리프팅(탄력)"),
            ("skinbooster_code", "스킨부스터"),
            ("volume_code", "볼꺼짐"),
        ]

        with self._session() as session:
            for property_name, feature_name in mappings:
                query = f"""
                MATCH (c:CombinedRule)
                MATCH (f:FeatureRule {{feature_name: $feature_name, age_group: c.age_group, code: c.{property_name}}})
                MERGE (c)-[:USES_CODE {{feature_name: $feature_name}}]->(f)
                """
                session.run(query, feature_name=feature_name)

    def debug_counts(self) -> Dict[str, int]:
        with self._session() as session:
            feature_count = session.run("MATCH (n:FeatureRule) RETURN count(n) AS c").single()["c"]
            # combined_count = session.run("MATCH (n:CombinedRule) RETURN count(n) AS c").single()["c"]
            # rel_count = session.run("MATCH ()-[r:USES_CODE]->() RETURN count(r) AS c").single()["c"]
            # chunk_count = session.run("MATCH (n:TreatmentChunk) RETURN count(n) AS c").single()["c"]
        return {
            "feature_rules": feature_count,
            # "combined_rules": combined_count,
            # "uses_code_relationships": rel_count,
            # "treatment_chunks": chunk_count,
        }


def ensure_required_env() -> None:
    required = {
        "OPENAI_API_KEY": os.getenv("OPENAI_API_KEY", ""),
        "NEO4J_URI": NEO4J_URI,
        "NEO4J_USERNAME": NEO4J_USERNAME,
        "NEO4J_PASSWORD": NEO4J_PASSWORD,
    }
    missing = [key for key, value in required.items() if not value]
    if missing:
        raise EnvironmentError(f"필수 환경변수가 없습니다: {', '.join(missing)}")


def validate_inputs(excel_path: str) -> str:
    resolved_path = Path(excel_path).expanduser().resolve()
    if not resolved_path.exists():
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {resolved_path}")
    if resolved_path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError(f"지원하지 않는 엑셀 파일 형식입니다: {resolved_path.name}")
    return str(resolved_path)


def main() -> None:
    ensure_required_env()
    excel_path = validate_inputs(EXCEL_PATH)

    feature_rules, combined_rules = load_rules_from_excel(excel_path)
    print(f"[1/6] 엑셀 파싱 완료: feature_rules={len(feature_rules)}, combined_rules={len(combined_rules)}")

    uploader = AuraUploader(
        uri=NEO4J_URI,
        username=NEO4J_USERNAME,
        password=NEO4J_PASSWORD,
        database=NEO4J_DATABASE,
    )

    try:
        uploader.verify_connectivity()
        print("[2/6] Neo4j 연결 확인 완료")

        uploader.wipe_existing()
        print("[2.5/6] 기존 FeatureRule/CombinedRule 노드 및 벡터 인덱스 제거 완료")

        embedder = Embedder(model=OPENAI_EMBED_MODEL, dimensions=OPENAI_EMBED_DIMENSIONS)
        feature_vectors = embedder.embed_texts([r.customer_desc for r in feature_rules], batch_size=EMBED_BATCH_SIZE)
        # combined_vectors = embedder.embed_texts([r.embedding_text for r in combined_rules], batch_size=EMBED_BATCH_SIZE)
        print("[3/6] 임베딩 생성 완료")

        uploader.setup_schema(OPENAI_EMBED_DIMENSIONS)
        print("[4/6] 제약조건/벡터 인덱스 생성 완료")

        uploader.upload_feature_rules(feature_rules, feature_vectors)
        # uploader.upload_combined_rules(combined_rules, combined_vectors)
        print("[5/6] 노드 업로드 완료")

        # uploader.create_relationships()
        # print("[6/6] 관계 생성 완료")

        counts = uploader.debug_counts()
        print("업로드 결과:")
        for key, value in counts.items():
            print(f"  - {key}: {value}")

    finally:
        uploader.close()


if __name__ == "__main__":
    main()
